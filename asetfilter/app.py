"""
AsetFilter - Main Flask Application

A web application for parsing, filtering, and exporting government asset data
from complex Excel files.
"""
import os
from datetime import datetime
from flask import Flask, render_template, request, redirect, url_for, flash, send_file, jsonify
from werkzeug.utils import secure_filename
from flask_wtf.csrf import CSRFProtect
import pandas as pd
from io import BytesIO

from config import Config
from models import db, Asset, UploadHistory
from forms import UploadForm, FilterForm
from parser import parse_excel_file, get_unique_values, get_status_options, get_luas_range

# Initialize CSRF protection
csrf = CSRFProtect()


def create_app(config_class=Config):
    """Application factory pattern"""
    app = Flask(__name__)
    app.config.from_object(config_class)
    
    # Initialize extensions
    db.init_app(app)
    csrf.init_app(app)
    config_class.init_app(app)
    
    # Create database tables
    with app.app_context():
        db.create_all()
    
    return app


app = create_app()


def allowed_file(filename):
    """Check if file extension is allowed"""
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']


def get_filter_options():
    """Get current filter options from database"""
    kecamatan_list = db.session.query(Asset.kecamatan).distinct().filter(
        Asset.kecamatan.isnot(None), 
        Asset.kecamatan != ''
    ).order_by(Asset.kecamatan).all()
    kecamatan_choices = [('', 'Semua Kecamatan')] + [(k[0], k[0]) for k in kecamatan_list]
    
    # Get satuan kerja options
    satuan_kerja_list = db.session.query(Asset.satuan_kerja).distinct().filter(
        Asset.satuan_kerja.isnot(None), 
        Asset.satuan_kerja != ''
    ).order_by(Asset.satuan_kerja).all()
    satuan_kerja_choices = [('', 'Semua Satuan Kerja')] + [(s[0], s[0]) for s in satuan_kerja_list]
    
    # Get unique status values (combined - for backward compatibility)
    status_list = db.session.query(Asset.status_combined).distinct().filter(
        Asset.status_combined.isnot(None),
        Asset.status_combined != ''
    ).all()
    
    all_statuses = set()
    for s in status_list:
        if s[0]:
            parts = s[0].split('|')
            for part in parts:
                cleaned = part.strip()
                if cleaned:
                    all_statuses.add(cleaned)
    
    status_choices = [(s, s) for s in sorted(all_statuses)]
    
    # Get separate status field options
    def get_unique_choices(field):
        values = db.session.query(field).distinct().filter(
            field.isnot(None), field != ''
        ).order_by(field).all()
        return [(v[0], v[0]) for v in values if v[0]]
    
    status_tanah_choices = get_unique_choices(Asset.status_tanah)
    pemetaan_choices = get_unique_choices(Asset.pemetaan)
    catatan_choices = get_unique_choices(Asset.catatan)
    k3_choices = get_unique_choices(Asset.k3)
    tanah_bangunan_choices = get_unique_choices(Asset.tanah_bangunan)
    asal_usul_choices = get_unique_choices(Asset.asal_usul)
    lain_lain_choices = get_unique_choices(Asset.lain_lain)
    alamat_choices = get_unique_choices(Asset.alamat)
    
    # Get luas range
    luas_range = db.session.query(
        db.func.min(Asset.luas),
        db.func.max(Asset.luas)
    ).first()
    
    return {
        'kecamatan_choices': kecamatan_choices,
        'satuan_kerja_choices': satuan_kerja_choices,
        'status_choices': status_choices,
        'status_tanah_choices': status_tanah_choices,
        'pemetaan_choices': pemetaan_choices,
        'catatan_choices': catatan_choices,
        'k3_choices': k3_choices,
        'tanah_bangunan_choices': tanah_bangunan_choices,
        'asal_usul_choices': asal_usul_choices,
        'lain_lain_choices': lain_lain_choices,
        'alamat_choices': alamat_choices,
        'min_luas': luas_range[0] or 0,
        'max_luas': luas_range[1] or 0
    }


def apply_filters(query, filters):
    """Apply filters to asset query"""
    from sqlalchemy import or_
    
    if filters.get('nama_asset'):
        search_term = f"%{filters['nama_asset']}%"
        query = query.filter(Asset.nama_asset.ilike(search_term))
    
    if filters.get('kecamatan'):
        query = query.filter(Asset.kecamatan == filters['kecamatan'])
        
    if filters.get('satuan_kerja'):
        query = query.filter(Asset.satuan_kerja == filters['satuan_kerja'])
        
    if filters.get('alamat'):
        search_term = f"%{filters['alamat']}%"
        query = query.filter(Asset.alamat.ilike(search_term))
    
    if filters.get('min_luas') is not None:
        query = query.filter(Asset.luas >= filters['min_luas'])
    
    if filters.get('max_luas') is not None:
        query = query.filter(Asset.luas <= filters['max_luas'])
    
    # Helper function to filter for blank/empty values
    def filter_blank_or_value(field, value):
        if value == '__BLANK__':
            # Filter for NULL or empty string
            return or_(field.is_(None), field == '')
        else:
            return field.ilike(f"%{value}%")
    
    # Separate status filters
    if filters.get('status_tanah'):
        query = query.filter(filter_blank_or_value(Asset.status_tanah, filters['status_tanah']))
    
    if filters.get('pemetaan'):
        query = query.filter(filter_blank_or_value(Asset.pemetaan, filters['pemetaan']))
    
    if filters.get('catatan'):
        query = query.filter(filter_blank_or_value(Asset.catatan, filters['catatan']))
    
    if filters.get('k3'):
        query = query.filter(filter_blank_or_value(Asset.k3, filters['k3']))
    
    if filters.get('tanah_bangunan'):
        query = query.filter(filter_blank_or_value(Asset.tanah_bangunan, filters['tanah_bangunan']))
    
    if filters.get('asal_usul'):
        query = query.filter(filter_blank_or_value(Asset.asal_usul, filters['asal_usul']))
    
    if filters.get('lain_lain'):
        query = query.filter(filter_blank_or_value(Asset.lain_lain, filters['lain_lain']))
    
    # Combined status filter (legacy support)
    if filters.get('status'):
        status_list = filters['status']
        if isinstance(status_list, str):
            status_list = [status_list]
        
        conditions = []
        for status in status_list:
            conditions.append(Asset.status_combined.ilike(f"%{status}%"))
        
        if conditions:
            query = query.filter(or_(*conditions))
    
    return query


# ============================================================================
# ROUTES
# ============================================================================

@app.route('/')
def index():
    """Home page with filter form and results table"""
    # Get filter options
    options = get_filter_options()
    
    # Initialize form
    form = FilterForm()
    form.kecamatan.choices = options['kecamatan_choices']
    form.satuan_kerja.choices = options['satuan_kerja_choices']
    form.status.choices = options['status_choices']
    
    # Get filter parameters from request
    filters = {
        'nama_asset': request.args.get('nama_asset', ''),
        'kecamatan': request.args.get('kecamatan', ''),
        'satuan_kerja': request.args.get('satuan_kerja', ''),
        'alamat': request.args.get('alamat', ''),
        'min_luas': request.args.get('min_luas', type=float),
        'max_luas': request.args.get('max_luas', type=float),
        'status': request.args.getlist('status'),
        # Separate status filters
        'status_tanah': request.args.get('status_tanah', ''),
        'pemetaan': request.args.get('pemetaan', ''),
        'catatan': request.args.get('catatan', ''),
        'k3': request.args.get('k3', ''),
        'tanah_bangunan': request.args.get('tanah_bangunan', ''),
        'asal_usul': request.args.get('asal_usul', ''),
        'lain_lain': request.args.get('lain_lain', '')
    }
    
    # Pre-fill form with filter values (basic form fields only)
    form.nama_asset.data = filters['nama_asset']
    form.kecamatan.data = filters['kecamatan']
    form.satuan_kerja.data = filters['satuan_kerja']
    form.alamat.data = filters['alamat']
    form.min_luas.data = filters['min_luas']
    form.max_luas.data = filters['max_luas']
    form.status.data = filters['status']
    
    # Get pagination parameters
    page = request.args.get('page', 1, type=int)
    per_page = app.config.get('ROWS_PER_PAGE', 20)
    
    # Get sort parameters
    sort_by = request.args.get('sort', 'id')
    sort_order = request.args.get('order', 'asc')
    
    # Build query
    query = Asset.query
    
    # Apply filters
    query = apply_filters(query, filters)
    
    # Get total count before pagination
    total_count = Asset.query.count()
    filtered_count = query.count()
    
    # Apply sorting
    sort_column = getattr(Asset, sort_by, Asset.id)
    if sort_order == 'desc':
        query = query.order_by(sort_column.desc())
    else:
        query = query.order_by(sort_column.asc())
    
    # Paginate
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)
    assets = pagination.items
    
    return render_template('index.html',
                          form=form,
                          assets=assets,
                          pagination=pagination,
                          total_count=total_count,
                          filtered_count=filtered_count,
                          filters=filters,
                          sort_by=sort_by,
                          sort_order=sort_order,
                          options=options)


@app.route('/upload', methods=['GET', 'POST'])
def upload():
    """Upload page for Excel files"""
    form = UploadForm()
    
    if form.validate_on_submit():
        file = form.file.data
        
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            
            # Create upload record
            upload_record = UploadHistory(
                filename=filename,
                status='processing'
            )
            db.session.add(upload_record)
            db.session.commit()
            
            try:
                # Save file
                file.save(filepath)
                
                # Parse Excel file
                df, stats = parse_excel_file(filepath)
                
                if df.empty:
                    raise ValueError("No data could be extracted from the file")
                
                # Clear existing data
                Asset.query.delete()
                
                # Insert new data
                for _, row in df.iterrows():
                    asset = Asset(
                        no_kib=row.get('no_kib'),
                        no_urut=row.get('no_urut'),
                        kode_lokasi=row.get('kode_lokasi'),
                        kode_aset=row.get('kode_aset'),
                        satuan_kerja=row.get('satuan_kerja'),
                        nama_asset=row.get('nama_asset'),
                        nomor=row.get('nomor'),
                        luas=row.get('luas'),
                        tahun=row.get('tahun'),
                        kecamatan=row.get('kecamatan'),
                        alamat=row.get('alamat'),
                        status_tanah=row.get('status_tanah'),
                        catatan=row.get('catatan'),
                        k3=row.get('k3'),
                        pemetaan=row.get('pemetaan'),
                        tanah_bangunan=row.get('tanah_bangunan'),
                        status_combined=row.get('status_combined'),
                        nilai_harga=row.get('nilai_harga'),
                        asal_usul=row.get('asal_usul'),
                        penggunaan=row.get('penggunaan'),
                        jumlah_bidang=row.get('jumlah_bidang'),
                        keterangan=row.get('keterangan'),
                        lain_lain=row.get('lain_lain')
                    )
                    db.session.add(asset)
                
                db.session.commit()
                
                # Update upload record
                upload_record.status = 'success'
                upload_record.records_count = stats['valid_rows']
                db.session.commit()
                
                flash(f'File berhasil diupload! {stats["valid_rows"]} data berhasil diproses.', 'success')
                
                # Get sample data for preview
                sample_assets = Asset.query.limit(10).all()
                
                return render_template('upload.html', 
                                      form=form, 
                                      success=True,
                                      stats=stats,
                                      sample_assets=sample_assets)
                
            except Exception as e:
                db.session.rollback()
                upload_record.status = 'failed'
                upload_record.error_message = str(e)
                db.session.commit()
                
                flash(f'Error: {str(e)}', 'error')
                
            finally:
                # Clean up uploaded file
                if os.path.exists(filepath):
                    os.remove(filepath)
        else:
            flash('File tidak valid. Hanya file .xls dan .xlsx yang diperbolehkan.', 'error')
    
    # Get last upload info
    last_upload = UploadHistory.query.order_by(UploadHistory.uploaded_at.desc()).first()
    current_count = Asset.query.count()
    
    return render_template('upload.html', 
                          form=form, 
                          last_upload=last_upload,
                          current_count=current_count)


@app.route('/filter', methods=['POST'])
def filter_ajax():
    """AJAX endpoint for filtering assets"""
    try:
        filters = {
            'nama_asset': request.form.get('nama_asset', ''),
            'kecamatan': request.form.get('kecamatan', ''),
            'satuan_kerja': request.form.get('satuan_kerja', ''),
            'alamat': request.form.get('alamat', ''),
            'min_luas': request.form.get('min_luas', type=float),
            'max_luas': request.form.get('max_luas', type=float),
            'status': request.form.getlist('status')
        }
        
        page = request.form.get('page', 1, type=int)
        per_page = app.config.get('ROWS_PER_PAGE', 20)
        sort_by = request.form.get('sort', 'id')
        sort_order = request.form.get('order', 'asc')
        
        query = Asset.query
        query = apply_filters(query, filters)
        
        total_count = Asset.query.count()
        filtered_count = query.count()
        
        sort_column = getattr(Asset, sort_by, Asset.id)
        if sort_order == 'desc':
            query = query.order_by(sort_column.desc())
        else:
            query = query.order_by(sort_column.asc())
        
        pagination = query.paginate(page=page, per_page=per_page, error_out=False)
        assets = [asset.to_dict() for asset in pagination.items]
        
        return jsonify({
            'success': True,
            'assets': assets,
            'total_count': total_count,
            'filtered_count': filtered_count,
            'page': page,
            'total_pages': pagination.pages,
            'has_next': pagination.has_next,
            'has_prev': pagination.has_prev
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})





@app.route('/export-excel')
def export_excel():
    """Export filtered data to Excel with multi-row header format matching Export_Filter.xlsx"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    try:
        filters = {
            'nama_asset': request.args.get('nama_asset', ''),
            'kecamatan': request.args.get('kecamatan', ''),
            'satuan_kerja': request.args.get('satuan_kerja', ''),
            'alamat': request.args.get('alamat', ''),
            'min_luas': request.args.get('min_luas', type=float),
            'max_luas': request.args.get('max_luas', type=float),
            'status': request.args.getlist('status'),
            
            # Detailed status filters
            'status_tanah': request.args.get('status_tanah', ''),
            'pemetaan': request.args.get('pemetaan', ''),
            'catatan': request.args.get('catatan', ''),
            'k3': request.args.get('k3', ''),
            'tanah_bangunan': request.args.get('tanah_bangunan', ''),
            'asal_usul': request.args.get('asal_usul', ''),
            'lain_lain': request.args.get('lain_lain', '')
        }
        
        query = Asset.query
        query = apply_filters(query, filters)
        assets = query.all()
        
        if not assets:
            flash('Tidak ada data untuk diexport.', 'warning')
            return redirect(url_for('index'))
        
        # Create workbook with custom header format
        wb = Workbook()
        ws = wb.active
        ws.title = 'Data Aset'
        
        # Define styles
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        header_font = Font(bold=True)
        red_font = Font(bold=True, color='FF0000')
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Define column structure matching Export_Filter.xlsx
        # Row 1: Empty (for spacing like the original)
        # Row 2-3: Headers with merged cells
        # Data starts from row 4
        
        # Column mapping: (col_index, header_row1, header_row2, is_red, db_field)
        columns = [
            (1, 'NO. KIB 2023', None, False, 'no_kib'),
            (2, 'No.', None, False, 'no_urut'),
            (3, 'Kode Lokasi', None, False, 'kode_lokasi'),
            (4, 'Satuan Kerja', None, False, 'satuan_kerja'),
            (5, 'Jenis Barang / Nama Barang', None, False, 'nama_asset'),
            (6, 'Nomor', 'Kd Barang', False, None),  # Sub-header only
            (7, None, 'Reg', False, 'nomor'),
            (8, 'Luas (m2)', None, False, 'luas'),
            (9, 'Tahun', 'Pengadaan', False, 'tahun'),
            (10, None, 'Letak / Alamat', False, 'alamat'),
            (11, 'Status Tanah', 'Hak', False, 'status_tanah'),
            (12, 'Sertifikat', 'Tanggal', False, None),  # Sub-header only
            (13, None, 'No.', False, None),  # Sub-header only
            (14, 'Penggunaan', None, False, 'nama_asset'),  # Using nama_asset as per model
            (15, 'Asal Usul', None, False, 'asal_usul'),
            (16, 'Nilai / Harga', None, False, 'nilai_harga'),
            (17, 'Keterangan', None, False, 'keterangan'),
            (18, 'Kode Aset', None, False, 'kode_aset'),
            (19, 'JUMLAH BIDANG', None, True, 'jumlah_bidang'),
            (20, 'KECAMATAN', None, True, 'kecamatan'),
            (21, 'PEMETAAN ASET TANAH', None, True, 'pemetaan'),
            (22, 'CATATAN (TERMANFAATKAN/TERLANTAR)', None, True, 'catatan'),
            (23, 'K3 (MILIK WARGA/ADA KLAIM, TKD, DLL)', None, True, 'k3'),
            (24, 'TANAH (BANGUNAN/TANAH KOSONG)', None, True, 'tanah_bangunan'),
            (25, 'LAIN-LAIN', None, False, 'lain_lain'),
        ]
        
        # Write headers - Row 3 (main headers) and Row 4 (sub-headers)
        header_row1 = 3
        header_row2 = 4
        
        for col_idx, header1, header2, is_red, _ in columns:
            cell1 = ws.cell(row=header_row1, column=col_idx)
            cell2 = ws.cell(row=header_row2, column=col_idx)
            
            if header1:
                cell1.value = header1
                cell1.font = red_font if is_red else header_font
                cell1.alignment = center_align
                cell1.border = thin_border
                
                # If no sub-header, merge cells vertically
                if not header2:
                    ws.merge_cells(start_row=header_row1, start_column=col_idx, 
                                   end_row=header_row2, end_column=col_idx)
            
            if header2:
                cell2.value = header2
                cell2.font = header_font
                cell2.alignment = center_align
                cell2.border = thin_border
            elif not header1:
                # Part of a merged cell from above, still apply border
                cell2.border = thin_border
        
        # Merge "Nomor" header across columns 6-7
        ws.merge_cells(start_row=header_row1, start_column=6, end_row=header_row1, end_column=7)
        
        # Merge "Tahun" header across columns 9-10
        ws.merge_cells(start_row=header_row1, start_column=9, end_row=header_row1, end_column=10)
        
        # Merge "Status Tanah" header across columns 11
        # Merge "Sertifikat" header across columns 12-13
        ws.merge_cells(start_row=header_row1, start_column=12, end_row=header_row1, end_column=13)
        
        # Write data starting from row 5
        data_start_row = 5
        for row_idx, asset in enumerate(assets, start=data_start_row):
            for col_idx, _, _, _, db_field in columns:
                if db_field:
                    value = getattr(asset, db_field, None)
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = thin_border
                else:
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.border = thin_border
        
        # Adjust column widths
        column_widths = {
            1: 12, 2: 5, 3: 12, 4: 25, 5: 30, 6: 10, 7: 8, 8: 10,
            9: 10, 10: 25, 11: 12, 12: 10, 13: 8, 14: 20, 15: 12,
            16: 15, 17: 20, 18: 15, 19: 12, 20: 15, 21: 18, 22: 30,
            23: 30, 24: 25, 25: 15
        }
        for col_idx, width in column_widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'aset_filter_export_{timestamp}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        flash(f'Error exporting Excel: {str(e)}', 'error')
        return redirect(url_for('index'))


@app.route('/clear-data', methods=['POST'])
def clear_data():
    """Clear all asset data"""
    try:
        count = Asset.query.count()
        Asset.query.delete()
        db.session.commit()
        
        flash(f'{count} data berhasil dihapus.', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')
    
    return redirect(url_for('index'))


@app.route('/api/stats')
def api_stats():
    """API endpoint for dashboard statistics"""
    total = Asset.query.count()
    
    # Count by kecamatan
    by_kecamatan = db.session.query(
        Asset.kecamatan, 
        db.func.count(Asset.id)
    ).filter(Asset.kecamatan.isnot(None)).group_by(Asset.kecamatan).all()
    
    # Count by satuan kerja
    by_satuan_kerja = db.session.query(
        Asset.satuan_kerja,
        db.func.count(Asset.id)
    ).filter(Asset.satuan_kerja.isnot(None)).group_by(Asset.satuan_kerja).limit(10).all()
    
    # Total luas
    total_luas = db.session.query(db.func.sum(Asset.luas)).scalar() or 0
    
    return jsonify({
        'total': total,
        'total_luas': float(total_luas),
        'by_kecamatan': dict(by_kecamatan),
        'by_satuan_kerja': dict(by_satuan_kerja)
    })


# ============================================================================
# ERROR HANDLERS
# ============================================================================

@app.errorhandler(413)
def too_large(e):
    flash('File terlalu besar. Maksimum 10MB.', 'error')
    return redirect(url_for('upload'))


@app.errorhandler(404)
def not_found(e):
    return render_template('404.html'), 404


@app.errorhandler(500)
def server_error(e):
    return render_template('500.html'), 500


# ============================================================================
# MAIN
# ============================================================================

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
